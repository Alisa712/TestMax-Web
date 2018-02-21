import json
from openpyxl import load_workbook

workbook = load_workbook('./MC QuestionSet.xlsx')
f = open('./parsed.js','wb')
fir_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(fir_sheet)
f.write("var data = ")
arr = []

for row in worksheet.iter_rows():
	qid = row[0].value
	des = row[1].value
	cat = row[2].value
	cho = []
	ans = -1
	for idx in range(3,8):
		cho.append(row[idx].value)
		# Color Green
		if row[idx].fill.start_color.index == 13:
			ans = idx - 2
	final_data = {'Id':qid,'Description':des, 'Category': cat, 'Choices':cho, 'Ans':ans}
	if ans != -1:
		arr.append(final_data)
		# f.write(json.dumps(final_data) +',\n')
	else:
		print "Error in line"
f.write(json.dumps(arr))
f.close()